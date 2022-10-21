import musicbrainzngs as mb
import openpyxl
import excel_table
import json
import sys
from progress.bar import IncrementalBar
from datetime import datetime

def init_config():
    try:
        config_file = open("data_config.json", "r")
        global CONFIG
        CONFIG = json.load(config_file)
        return True
    except IOError as e:
        print(e)
        return False

def album_has_keywords(album_dict):
    if "tag-list" not in album_dict:
        return False
    return True

def add_to_distribution(key, distr_dict: dict):
    if key not in distr_dict:
        distr_dict[key] = 0
    distr_dict[key] += 1

def collect_album_tag_data(album, genre: str):
    if not album_has_keywords(album):
        return None
    
    total_tags = 0
    tags_containing_genre = 0

    for tag in album["tag-list"]:
        tag_upvotes = int(tag["count"])

        if tag_upvotes < 1:
            continue

        total_tags += tag_upvotes

        if genre in tag["name"]:
            tags_containing_genre += tag_upvotes

    
    if total_tags and 0.5 < (tags_containing_genre / total_tags):
        return {"total_tags": total_tags, "tags_containing_genre": tags_containing_genre}
    return None
        

def only_big_letters(char: chr): #change name of function
    char_to_dec = ord(char)
    
    if char_to_dec in range(65, 91):
        return char
    if char_to_dec in range(97, 123):
        return chr(char_to_dec - 32)
    return "other"

def collect_bias_distr_data(album):
    distribution_data = {
        "artist_first_chr": None,
        "title_first_chr":  None,
        "release_year":     None
    }

    if "first-release-date" in album:
        distribution_data["release_year"] = int(album["first-release-date"].split('-')[0])
    if "artist-credit-phrase" in album:
        distribution_data["artist_first_chr"] = only_big_letters(album["artist-credit-phrase"][0])
    if "title" in album:
        distribution_data["title_first_chr"] = only_big_letters(album["title"][0])

    return distribution_data

def get_data_for_genres(genres, check_for_album_covers):
    mb.set_useragent("test", "0.7.1")

    data_statistics = {"genres": []}

    artist_first_chr_distr = {}
    title_first_chr_distr =  {}
    release_year_distr =     {}

    for i in range(0, len(genres)):

        num_albums_to_collect = int(sys.argv[1])
        offset                = 0
        
        tags_containing_genre = 0
        total_tags            = 0

        while 0 < num_albums_to_collect:
            limit = 100 if 100 <= num_albums_to_collect else num_albums_to_collect
            result = mb.search_release_groups(query='', limit=limit, offset=offset, primarytype="Album", tag=genres[i])

            if len(result["release-group-list"]) == 0:
                print("no more albums to collect")
                break

            offset += len(result["release-group-list"])
        
            albums = result["release-group-list"]

            num_collected_albums = 0

            for album in albums:

                tag_data = collect_album_tag_data(album, genres[i])

                if tag_data is None:
                    continue
                tags_containing_genre += tag_data["tags_containing_genre"]
                total_tags += tag_data["total_tags"]
                
                num_collected_albums += 1

                distr_data = collect_bias_distr_data(album)

                if distr_data["artist_first_chr"] != None:
                    add_to_distribution(distr_data["artist_first_chr"], artist_first_chr_distr)
                if distr_data["title_first_chr"] != None:
                    add_to_distribution(distr_data["title_first_chr"], title_first_chr_distr)
                if distr_data["release_year"] != None:
                    add_to_distribution(distr_data["release_year"], release_year_distr)

            
            num_albums_to_collect -= num_collected_albums

        data_statistics["genres"].append({"genre": genres[i], "total_tags": total_tags, "genre_tags": tags_containing_genre})

    data_statistics["artist_first_chr_distr"] = artist_first_chr_distr
    data_statistics["title_first_chr_distr"]  = title_first_chr_distr
    data_statistics["release_year_distr"]     = release_year_distr

    return data_statistics


def input_argument_formatting_control():
    if len(sys.argv) < 3:
        print("Too few arguments: retrieve_covers.py <num_of_covers> <output_file>")
        exit(-1)
    
    if not sys.argv[1].isdigit():
        print("Wrong type: <num_of_covers> must be integer")
        exit(-1)

if __name__ == "__main__":
    
    input_argument_formatting_control()

    if not init_config():
        exit(-1)

    genres = CONFIG["album_genres"]

    #retrieve the statistics of album covers
    statistics = get_data_for_genres(genres, CONFIG["check_for_album_covers"])

    print(statistics)

    #create and format an excel document based on the statistics
    workbook = openpyxl.Workbook()
    del workbook[workbook.sheetnames[0]]

    first_chr_artist_sheet = workbook.create_sheet("First char artist")
    first_chr_title_sheet  = workbook.create_sheet("First char title")
    year_distr_sheet       = workbook.create_sheet("Release year distribution")
    stats_sheet            = workbook.create_sheet("album cover statistics")

    excel_table.add_distribution_to_sheet(year_distr_sheet, sorted(statistics["release_year_distr"].items()), "year", "number of albums")
    excel_table.add_distribution_to_sheet(first_chr_title_sheet, sorted(statistics["title_first_chr_distr"].items()), "char", "number of albums")
    excel_table.add_distribution_to_sheet(first_chr_artist_sheet, sorted(statistics["artist_first_chr_distr"].items()), "char", "number of albums")

    excel_table.init_stats_sheet_titles(stats_sheet)

    row = 2
    for i in range(len(statistics["genres"])):
        data = statistics["genres"][i]
        col = 1

        for value in data.values():
            stats_sheet.cell(row, col).value = value
            col += 1

        stats_sheet.cell(row, col).value = int(sys.argv[1]) #number of albums   
        
        row += 1


    workbook.save(str(sys.argv[2]))
