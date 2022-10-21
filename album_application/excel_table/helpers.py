import openpyxl

def init_stats_sheet_titles(sheet):
    if type(sheet) is not openpyxl.worksheet.worksheet.Worksheet:
        return False
    
    sheet["A1"] = "genre"
    sheet["B1"] = "total tags album"
    sheet["C1"] = "tags containing search genre"
    sheet["D1"] = "total albums"
    return True

def add_distribution_to_sheet(sheet, distr_dict: dict, key_title: str, number_title: str):
    sheet.cell(1, 1).value = key_title
    sheet.cell(1, 2).value = number_title
    current_row = 2

    for key, value in distr_dict:
        sheet.cell(current_row, 1).value = key
        sheet.cell(current_row, 2).value = value
        current_row += 1